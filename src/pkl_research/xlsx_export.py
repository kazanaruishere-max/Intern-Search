"""Export shortlist ke Excel (xlsx) dengan styling agar mudah dibaca."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F7FB")
THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

FILL_FIT_HIGH = PatternFill("solid", fgColor="C6EFCE")
FILL_FIT_MID = PatternFill("solid", fgColor="E2EFDA")
FILL_FIT_LOW = PatternFill("solid", fgColor="FFF2CC")
FILL_AI = PatternFill("solid", fgColor="FCE4D6")
FILL_RATING = PatternFill("solid", fgColor="DDEBF7")
FILL_NEAR = PatternFill("solid", fgColor="E2EFDA")
FILL_MID = PatternFill("solid", fgColor="FFF2CC")
FILL_FAR = PatternFill("solid", fgColor="FCE4D6")
FILL_HEADER_SUB = PatternFill("solid", fgColor="D6E4F0")

LEGEND = [
    ("Fit ≥ 90", FILL_FIT_HIGH),
    ("Fit 80–89", FILL_FIT_MID),
    ("Fit 70–79", FILL_FIT_LOW),
    ("Ada unsur AI (dari website)", FILL_AI),
    ("Rating ≥ 4.9", FILL_RATING),
    ("Jarak ≤ 3 km", FILL_NEAR),
    ("Jarak 3–5 km", FILL_MID),
    ("Jarak > 5 km", FILL_FAR),
]


def _style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def _fit_fill(fit: float) -> PatternFill:
    if fit >= 90:
        return FILL_FIT_HIGH
    if fit >= 80:
        return FILL_FIT_MID
    return FILL_FIT_LOW


def _distance_fill(km: float | None) -> PatternFill:
    if km is None:
        return PatternFill()
    if km <= 3:
        return FILL_NEAR
    if km <= 5:
        return FILL_MID
    return FILL_FAR


def _set_widths(ws, widths: list[float]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def export_shortlist_xlsx(
    items: list[tuple[object, float, bool]],
    profiles_by_id: dict[int, object],
    drafts_by_company: dict[str, dict[str, str]],
    path: str | Path,
    notes_by_id: dict[int, str] | None = None,
    status_by_id: dict[int, str] | None = None,
) -> None:
    """Buat shortlist.xlsx: 3 sheet (Shortlist / Profil & Tentang / Draft Pesan)."""
    notes_by_id = notes_by_id or {}
    status_by_id = status_by_id or {}
    wb = Workbook()

    # ---------------- Sheet 1: Shortlist ----------------
    ws = wb.active
    ws.title = "Shortlist"
    headers = [
        "No", "Nama Perusahaan", "Fit-CV", "AI", "Rating", "Ulasan", "Jarak (km)",
        "Kategori", "Role", "Sektor", "Alamat", "Telepon", "WhatsApp", "Website",
        "Email", "Halaman Karir", "LinkedIn", "Fokus", "Health", "Red Flags",
        "Green Flags", "Catatan", "Status",
    ]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="SHORTLIST CV-MATCH — Perusahaan IT Jakarta Selatan")
    title_cell.font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    legend_cell = ws.cell(
        row=2, column=1,
        value="Legenda warna:  " + "   |   ".join(label for label, _ in LEGEND),
    )
    legend_cell.font = Font(size=9, italic=True, color="404040")

    header_row = 4
    for col, name in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=name)
    _style_header(ws, header_row, len(headers))

    for idx, (company, fit, ai) in enumerate(items, start=1):
        row = header_row + idx
        profile = profiles_by_id.get(company.id)
        health_label = ""
        red_summary = ""
        green_summary = ""
        greens = 0
        reds = 0
        if profile:
            if getattr(profile, "ai_focus", False):
                greens += 1
            if getattr(profile, "career_page_found", False):
                greens += 1
            if getattr(profile, "emails", None):
                greens += 1
            if getattr(profile, "site_title", None) and getattr(profile, "meta_description", None):
                greens += 1
            if getattr(profile, "fetch_status", "") == "failed":
                reds += 3
            if not getattr(profile, "emails", None):
                reds += 1
        ai_text = ""
        if profile and getattr(profile, "ai_focus", False):
            ai_text = "YA — " + " / ".join(getattr(profile, "ai_subfields", []))
        elif ai:
            ai_text = "YA (tag role)"
        elif profile and getattr(profile, "fetch_status", "") == "failed":
            ai_text = "site gagal"

        hs = max(0, min(100, 50 + greens * 10 - reds * 15))
        if hs >= 70:
            health_label = f"🟢 {hs}"
        elif hs >= 40:
            health_label = f"🟡 {hs}"
        else:
            health_label = f"🔴 {hs}"

        values = [
            idx,                                                    # No
            company.name,                                           # Nama
            fit,                                                    # Fit-CV
            ai_text,                                                # AI
            company.rating if company.rating is not None else "",   # Rating
            company.review_count if company.review_count is not None else "",  # Ulasan
            company.distance_km if company.distance_km is not None else "",    # Jarak
            company.category or "",                                 # Kategori
            ", ".join(company.role_fit),                            # Role
            company.sector or "",                                   # Sektor
            company.address or "",                                  # Alamat
            company.phone or "",                                    # Telepon
            (profile.whatsapp or "") if profile and profile.whatsapp else "",  # WhatsApp
            company.website or "",                                  # Website
            ", ".join(profile.emails) if profile and profile.emails else "",   # Email
            "ya" if profile and profile.career_page_found else "",  # Karir
            (profile.linkedin_url or "") if profile and profile.linkedin_url else "",  # LinkedIn
            (profile.core_focus or "") if profile else "",          # Fokus
            health_label,                                           # Health
            red_summary,                                            # Red Flags
            green_summary,                                          # Green Flags
            notes_by_id.get(company.id, ""),                        # Catatan
            status_by_id.get(company.id, ""),                       # Status
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = WRAP
            if idx % 2 == 0:
                cell.fill = ZEBRA_FILL
        ws.cell(row=row, column=3).fill = _fit_fill(fit)
        if ai_text.startswith("YA"):
            ws.cell(row=row, column=4).fill = FILL_AI
        if company.rating is not None and company.rating >= 4.9:
            ws.cell(row=row, column=5).fill = FILL_RATING
        ws.cell(row=row, column=7).fill = _distance_fill(company.distance_km)

    _set_widths(ws, [5, 34, 8, 22, 8, 8, 9, 20, 14, 9, 40, 16, 20, 26, 24, 10, 38, 45, 12, 40, 40, 60, 10])
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(items)}"
    for col in (6, 10):
        for r in range(header_row + 1, header_row + len(items) + 1):
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    # ---------------- Sheet 2: Profil & Tentang ----------------
    ws2 = wb.create_sheet("Profil & Tentang")
    headers2 = ["No", "Nama Perusahaan", "Fokus", "Tentang (about)", "Layanan (services)", "AI Bukti"]
    for col, name in enumerate(headers2, start=1):
        ws2.cell(row=1, column=col, value=name)
    _style_header(ws2, 1, len(headers2))
    for idx, (company, fit, ai) in enumerate(items, start=1):
        row = idx + 1
        profile = profiles_by_id.get(company.id)
        evidence = ""
        if profile and getattr(profile, "ai_focus", False):
            evidence = "\n".join(f'• "{e}"' for e in (getattr(profile, "ai_evidence", None) or [])[:3])
        values = [
            idx,
            company.name,
            (profile.core_focus or "") if profile else "",
            (profile.about_text or "") if profile else "",
            (profile.services_text or "") if profile else "",
            evidence,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.border = BORDER
            cell.alignment = WRAP
        ws2.row_dimensions[row].height = 90
    _set_widths(ws2, [5, 34, 45, 70, 60, 50])
    ws2.freeze_panes = "B2"

    # ---------------- Sheet 3: Draft Pesan ----------------
    ws3 = wb.create_sheet("Draft Pesan")
    headers3 = ["No", "Nama Perusahaan", "Varian", "Draft Pesan"]
    for col, name in enumerate(headers3, start=1):
        ws3.cell(row=1, column=col, value=name)
    _style_header(ws3, 1, len(headers3))
    r = 2
    for idx, (name, variants) in enumerate(drafts_by_company.items(), start=1):
        for variant, text in variants.items():
            ws3.cell(row=r, column=1, value=idx)
            ws3.cell(row=r, column=2, value=name)
            ws3.cell(row=r, column=3, value=variant)
            cell = ws3.cell(row=r, column=4, value=text)
            cell.alignment = WRAP
            for col in range(1, 5):
                ws3.cell(row=r, column=col).border = BORDER
            ws3.row_dimensions[r].height = 150
            r += 1
    _set_widths(ws3, [5, 34, 12, 100])
    ws3.freeze_panes = "A2"

    wb.save(path)
