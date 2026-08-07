from openpyxl import load_workbook

from pkl_research.models import Company, CompanyProfile
from pkl_research.xlsx_export import export_shortlist_xlsx


def make_items():
    c1 = Company(place_id="p1", name="PT Alpha", rating=4.9, review_count=120,
                 distance_km=2.0, category="Software house", role_fit=["ai"],
                 sector="swasta", address="Jl. Senopati, Jakarta Selatan",
                 phone="021-123", website="https://alpha.com", id=1)
    c2 = Company(place_id="p2", name="PT Beta", rating=4.5, review_count=40,
                 distance_km=6.0, category="Web agency", role_fit=["fullstack"],
                 sector="swasta", address="Kebayoran Baru, Jakarta Selatan",
                 website="https://beta.com", id=2)
    p1 = CompanyProfile(company_id=1, core_focus="AI dan machine learning",
                        about_text="Perusahaan AI.", ai_focus=True,
                        ai_subfields=["Machine Learning"], ai_evidence=["buat model ML"],
                        emails=["hr@alpha.com"], career_page_found=True)
    items = [(c1, 100.0, True), (c2, 73.0, False)]
    profiles = {1: p1}
    drafts = {"PT Alpha": {"formal": "Halo PT Alpha..."}}
    return items, profiles, drafts


def test_export_creates_three_sheets(tmp_path):
    path = tmp_path / "shortlist.xlsx"
    items, profiles, drafts = make_items()
    export_shortlist_xlsx(items, profiles, drafts, path)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Shortlist", "Profil & Tentang", "Draft Pesan"]
    wb.close()


def test_shortlist_header_and_rows(tmp_path):
    path = tmp_path / "shortlist.xlsx"
    items, profiles, drafts = make_items()
    export_shortlist_xlsx(items, profiles, drafts, path)
    wb = load_workbook(path)
    ws = wb["Shortlist"]
    assert ws.cell(row=4, column=2).value == "Nama Perusahaan"
    assert ws.cell(row=5, column=2).value == "PT Alpha"
    assert ws.cell(row=6, column=2).value == "PT Beta"
    assert "AI dan machine learning" in str(ws.cell(row=5, column=17).value)
    # fit score & jarak ikut terisi
    assert ws.cell(row=5, column=3).value == 100.0
    assert ws.cell(row=5, column=7).value == 2.0
    wb.close()


def test_profile_sheet_has_about(tmp_path):
    path = tmp_path / "shortlist.xlsx"
    items, profiles, drafts = make_items()
    export_shortlist_xlsx(items, profiles, drafts, path)
    wb = load_workbook(path)
    ws = wb["Profil & Tentang"]
    assert ws.cell(row=2, column=4).value == "Perusahaan AI."
    assert ws.cell(row=2, column=6).value.startswith('• "buat model ML"')
    wb.close()


def test_draft_sheet_has_variants(tmp_path):
    path = tmp_path / "shortlist.xlsx"
    items, profiles, drafts = make_items()
    export_shortlist_xlsx(items, profiles, drafts, path)
    wb = load_workbook(path)
    ws = wb["Draft Pesan"]
    assert ws.cell(row=2, column=2).value == "PT Alpha"
    assert ws.cell(row=2, column=3).value == "formal"
    assert ws.cell(row=2, column=4).value == "Halo PT Alpha..."
    wb.close()
